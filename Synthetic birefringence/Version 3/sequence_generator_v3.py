"""
sequence_generator_v3.py
========================
Generate the HWP/QWP angle sequence that synthesises a target retarder, and
write it to an Excel sheet for the acquisition software.

Uses synthetic_birefringence_v3, which gets signed-S3 handedness from the
grid-calibrated bfp_handedness_oracle (correct plate order/offsets/retardances
+ the ~0.64-wave microscope), instead of the old idealised oracle.

Handedness convention
----------------------
There is now a single handedness knob: bfp_handedness_oracle.HANDEDNESS_SIGN.
You do NOT set it here. For a linear focal target the global sign cancels from
the correction, so the default (+1) is fine; flip it in the oracle only if an
independent absolute reference (the dual-run with a known QWP) shows it
reversed. After changing it, delete any cached S3_signed.npy so it is rebuilt.
"""

from synthetic_birefringence_v3 import find_retarder_settings, SyntheticRetarder, plot_ellipse_grid

# ---------------------------------------------------------------------------
# Settings
# ---------------------------------------------------------------------------
cal_dir       = '.'         # folder with the five .npy files (+ optional cache)
delta_waves   = 0.25
fast_axis_deg = 110          # re-sweep this now that handedness is calibrated —
                            # the optimum should be broad, not a 5 deg spike.

phi_start     = 43.6
phi_step      = 14.85
n_steps       = 14

handedness_oracle = True    # False reproduces the old |S3|-only behaviour
use_cached_signed = True    # load S3_signed.npy if present (delete to rebuild)

import matplotlib.pyplot as plt
plt.close('all')

hwp, qwp, results = find_retarder_settings(
    cal_dir=cal_dir,
    delta_waves=delta_waves,
    fa_deg=fast_axis_deg,
    phi_start=phi_start,
    phi_step=phi_step,
    n_steps=n_steps,
    handedness_oracle=handedness_oracle,
    use_cached_signed=use_cached_signed,
)

import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

wb = Workbook()
ws = wb.active
ws.title = 'HWP QWP Angles'

header_font   = Font(name='Arial', bold=True, size=11)
header_fill   = PatternFill('solid', start_color='D9E1F2')
pred_fill     = PatternFill('solid', start_color='EAF3DE')
hand_fill     = PatternFill('solid', start_color='FCE4D6')
centre        = Alignment(horizontal='center')
header_border = Border(bottom=Side(style='medium', color='4472C4'))

headers = [
    ('C', 'HWP',     header_fill),
    ('D', 'QWP',     header_fill),
    ('E', 'psi pred', pred_fill),
    ('F', 'chi pred', pred_fill),
    ('G', 'DOLP',    pred_fill),
    ('H', 'd match', pred_fill),
    ('I', 'hand',    hand_fill),    # informational only; acquisition reads C & D
]

for col_letter, label, fill in headers:
    cell           = ws[f'{col_letter}3']
    cell.value     = label
    cell.font      = header_font
    cell.fill      = fill
    cell.alignment = centre
    cell.border    = header_border

value_font      = Font(name='Arial', size=11)
value_font_pred = Font(name='Arial', size=11, italic=True, color='3B6D11')
value_font_hand = Font(name='Arial', size=11, italic=True, color='9C4221')

for i, r in enumerate(results):
    row      = 4 + i
    is_last  = (i == len(results) - 1)
    bot_side = Side(style='thin', color='8EA9C1') if is_last else Side()

    for col_letter, key in (('C', 'hwp'), ('D', 'qwp')):
        cell           = ws[f'{col_letter}{row}']
        cell.value     = int(round(r[key]))
        cell.font      = value_font
        cell.alignment = centre
        cell.border    = Border(
            left=Side(style='thin', color='B8CCE4'),
            right=Side(style='thin', color='B8CCE4'),
            bottom=bot_side,
        )

    pred_values = [
        ('E', round(r['alpha_target'], 1)),
        ('F', round(r['chi_target'],   1)),
        ('G', round(r['dolp_target'],  3)),
        ('H', round(r['match_error'],  4)),
    ]
    for col_letter, value in pred_values:
        cell           = ws[f'{col_letter}{row}']
        cell.value     = value
        cell.font      = value_font_pred
        cell.alignment = centre
        cell.border    = Border(
            left=Side(style='thin', color='C0DD97'),
            right=Side(style='thin', color='C0DD97'),
            bottom=bot_side,
        )

    sign = r.get('gen_s3_sign', 0)
    hand = 'R' if sign > 0 else 'L' if sign < 0 else '-'
    cell           = ws[f'I{row}']
    cell.value     = hand
    cell.font      = value_font_hand
    cell.alignment = centre
    cell.border    = Border(
        left=Side(style='thin', color='F2C2A6'),
        right=Side(style='thin', color='F2C2A6'),
        bottom=bot_side,
    )

for col_letter, width in (('C',10),('D',10),('E',10),('F',10),('G',9),('H',10),('I',7)):
    ws.column_dimensions[col_letter].width = width

path     = os.path.abspath(os.getcwd())
fileName = f'delta_waves={delta_waves:.3f}_fast_axis={fast_axis_deg:2.0f}.xlsx'
wb.save(os.path.join(path, fileName))
print(f'Saved: {os.path.join(path, fileName)}')

sr = SyntheticRetarder(cal_dir=cal_dir)
sr.plot_results(results)

plot_ellipse_grid(results, title="BFP ellipses " + str(delta_waves) + 'λ' + ' fast axis = ' + str(fast_axis_deg) + ' deg')
